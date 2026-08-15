from __future__ import annotations

import re
from pathlib import Path
from typing import Any

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
WORKBOOK = (ROOT / "../Desert-Cat-完整项目测试用例-华创模板版.xlsx").resolve()
OUTPUT = ROOT / "cases.py"
ROUTE_RE = re.compile(r"路由/入口=([^；\n]+)")
COMPONENT_RE = re.compile(r"组件=([^；\n]+)")
PERMISSION_RE = re.compile(r"权限=([^；\n]+)")
PAGE_CONFIG = {
    "DC-UI-001": (".download-page", "download"),
    "DC-UI-002": (".login-page", "login"),
    "DC-UI-003": (".legal-page--agreement", "legal"),
    "DC-UI-004": (".legal-page", "legal"),
    "DC-UI-005": (".home-page", "home"),
    "DC-UI-006": (".feed-container", "feed"),
    "DC-UI-007": (".follow-page", "following"),
    "DC-UI-008": (".messages-page", "messages"),
    "DC-UI-009": (".mine-container", "mine"),
    "DC-UI-010": (".chat-layout", "chat"),
    "DC-UI-011": (".detail-page", "detail"),
    "DC-UI-012": (".search-page", "search"),
    "DC-UI-013": (".up-page", "user_profile"),
    "DC-UI-014": (".fl-page", "friends"),
    "DC-UI-015": (".af-page", "add_friend"),
    "DC-UI-016": (".tools-page", "tools"),
    "DC-UI-017": (".ranking-page", "ai_ranking"),
    "DC-UI-018": (".guide-page", "ai_guide"),
    "DC-UI-019": (".chat-wrap", "ai_chat"),
    "DC-UI-020": (".echarts-dashboard", "echarts"),
    "DC-UI-021": (".settings-page", "profile"),
    "DC-UI-022": (".feedback-page", "feedback"),
    "DC-UI-023": (".link-mgmt", "admin_links"),
    "DC-UI-024": (".text-studio", "publish"),
    "DC-UI-025": (".admin-announcements", "announcements"),
    "DC-UI-026": (".admin-permissions", "permissions"),
    "DC-UI-027": (".travel-page", "travel"),
    "DC-UI-028": (".music-view", "music"),
    "DC-UI-029": (".reg-page", "register_component"),
    "DC-UI-030": (".reset-page", "forgot_component"),
}


def match_value(pattern: re.Pattern[str], text: str) -> str | None:
    match = pattern.search(text)
    return match.group(1).strip() if match else None


def main() -> None:
    workbook = load_workbook(WORKBOOK, read_only=True, data_only=True)
    sheet = workbook["测试用例"]
    lines = ["from desertcat_ui.models import UiCase\n\n\nUI_CASES = [\n"]
    try:
        for row in sheet.iter_rows(min_row=2, values_only=True):
            case_id = str(row[0] or "").strip()
            if not case_id.startswith("DC-UI-"):
                continue
            test_input = str(row[5] or "").strip()
            expected = str(row[7] or "").strip()
            route = match_value(ROUTE_RE, test_input)
            if route == "组件":
                route = None
            component = match_value(COMPONENT_RE, test_input) or ""
            permission = match_value(PERMISSION_RE, test_input)
            if permission == "无特殊权限":
                permission = None
            elif permission:
                permission = re.split(r"[（(]", permission, maxsplit=1)[0].strip()
            root_selector, action_key = PAGE_CONFIG[case_id]
            requires_auth = "准备普通用户" in str(row[4] or "") or "未登录访问跳转" in expected
            lines.append(
                "    UiCase("
                f"case_id={case_id!r}, page_name={str(row[1] or '').strip()!r}, "
                f"title={str(row[2] or '').strip()!r}, priority={str(row[3] or '').strip()!r}, "
                f"route={route!r}, component={component!r}, requires_auth={requires_auth!r}, "
                f"permission={permission!r}, root_selector={root_selector!r}, action_key={action_key!r}),\n"
            )
    finally:
        workbook.close()
    lines.append("]\n")
    OUTPUT.write_text("".join(lines), encoding="utf-8")
    print("generated", len(lines) - 2, "UI cases")


if __name__ == "__main__":
    main()
